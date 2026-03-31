"""
Excel Dashboard Builder for LMS Dataset.
Creates a professional Excel workbook with:
- VLOOKUP formulas
- IF formulas
- Auto-filters
- Conditional formatting
- Pivot-style summary tables
- KPI cards
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.formatting.rule import (
    CellIsRule, DataBarRule, ColorScaleRule, FormulaRule
)
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
import os
import warnings
warnings.filterwarnings('ignore')


# ─── Style constants ───
DARK_BLUE = "1F4E79"
MED_BLUE = "2E75B6"
LIGHT_BLUE = "D6E4F0"
WHITE = "FFFFFF"
LIGHT_GRAY = "F2F2F2"
GREEN = "C6EFCE"
GREEN_FONT = "006100"
YELLOW = "FFEB9C"
YELLOW_FONT = "9C5700"
RED = "FFC7CE"
RED_FONT = "9C0006"

HEADER_FILL = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color=WHITE, size=11)
SUBHEADER_FILL = PatternFill(start_color=MED_BLUE, end_color=MED_BLUE, fill_type="solid")
SUBHEADER_FONT = Font(name="Calibri", bold=True, color=WHITE, size=10)
TITLE_FONT = Font(name="Calibri", bold=True, size=14, color=DARK_BLUE)
KPI_VALUE_FONT = Font(name="Calibri", bold=True, size=20, color=DARK_BLUE)
KPI_LABEL_FONT = Font(name="Calibri", size=10, color="666666")
NORMAL_FONT = Font(name="Calibri", size=10)
BOLD_FONT = Font(name="Calibri", bold=True, size=10)
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
ALT_ROW_FILL = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")
KPI_BORDER = Border(
    left=Side(style="medium", color=MED_BLUE),
    right=Side(style="medium", color=MED_BLUE),
    top=Side(style="medium", color=MED_BLUE),
    bottom=Side(style="medium", color=MED_BLUE),
)


def style_header_row(ws, row, max_col):
    """Apply header styling to a row."""
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def style_data_rows(ws, start_row, end_row, max_col):
    """Apply alternating row styling."""
    for row in range(start_row, end_row + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = NORMAL_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = THIN_BORDER
            if (row - start_row) % 2 == 1:
                cell.fill = ALT_ROW_FILL


def auto_fit_columns(ws, min_width=10, max_width=30):
    """Auto-fit column widths."""
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted = min(max(max_length + 2, min_width), max_width)
        ws.column_dimensions[col_letter].width = adjusted


# ═══════════════════════════════════════════════════
# SHEET 1: Training Data (with auto-filters)
# ═══════════════════════════════════════════════════
def create_data_sheet(wb, df):
    """Create the main data sheet with filters and conditional formatting."""
    ws = wb.active
    ws.title = "Training Data"
    ws.sheet_properties.tabColor = "1F4E79"

    # Title
    ws.merge_cells("A1:M1")
    ws["A1"] = "LMS TRAINING DATA — MASTER TRACKER"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="center")

    # Subtitle
    ws.merge_cells("A2:M2")
    ws["A2"] = f"Total Records: {len(df)} | Report Date: 31-Mar-2026"
    ws["A2"].font = Font(name="Calibri", italic=True, size=10, color="666666")
    ws["A2"].alignment = Alignment(horizontal="center")

    # Column headers at row 4
    headers = [
        "Employee ID", "Employee Name", "Department", "Course Name",
        "Assigned Date", "Due Date", "Completion Date", "Status",
        "Completion Days", "Is Overdue", "Month", "Days Allowed",
        "On Time"
    ]
    for col, header in enumerate(headers, 1):
        ws.cell(row=4, column=col, value=header)
    style_header_row(ws, 4, len(headers))

    # Data rows
    for i, (_, row) in enumerate(df.iterrows()):
        r = i + 5
        ws.cell(row=r, column=1, value=row['employee_id'])
        ws.cell(row=r, column=2, value=row['employee_name'])
        ws.cell(row=r, column=3, value=row['department'])
        ws.cell(row=r, column=4, value=row['course_name'])
        ws.cell(row=r, column=5, value=str(row['assigned_date'])[:10] if pd.notna(row['assigned_date']) else "")
        ws.cell(row=r, column=6, value=str(row['due_date'])[:10] if pd.notna(row['due_date']) else "")
        ws.cell(row=r, column=7, value=str(row['completion_date'])[:10] if pd.notna(row['completion_date']) else "")
        ws.cell(row=r, column=8, value=row['status'])
        ws.cell(row=r, column=9, value=row['completion_time_days'] if pd.notna(row['completion_time_days']) else "")
        ws.cell(row=r, column=10, value="Yes" if row['is_overdue'] else "No")
        ws.cell(row=r, column=11, value=row['assigned_month'])
        ws.cell(row=r, column=12, value=row['days_allowed'])
        ws.cell(row=r, column=13, value="Yes" if row['completed_on_time'] else "No")

    last_row = len(df) + 4
    style_data_rows(ws, 5, last_row, len(headers))

    # ─── AUTO-FILTERS ───
    ws.auto_filter.ref = f"A4:M{last_row}"

    # ─── CONDITIONAL FORMATTING ───
    # Status column (H): Green for Completed, Yellow for Pending
    ws.conditional_formatting.add(
        f"H5:H{last_row}",
        CellIsRule(operator="equal", formula=['"Completed"'],
                   fill=PatternFill(bgColor=GREEN), font=Font(color=GREEN_FONT))
    )
    ws.conditional_formatting.add(
        f"H5:H{last_row}",
        CellIsRule(operator="equal", formula=['"Pending"'],
                   fill=PatternFill(bgColor=YELLOW), font=Font(color=YELLOW_FONT))
    )

    # Overdue column (J): Red for Yes
    ws.conditional_formatting.add(
        f"J5:J{last_row}",
        CellIsRule(operator="equal", formula=['"Yes"'],
                   fill=PatternFill(bgColor=RED), font=Font(color=RED_FONT))
    )

    # Completion days - data bars
    ws.conditional_formatting.add(
        f"I5:I{last_row}",
        DataBarRule(start_type="min", end_type="max",
                    color="2E75B6", showValue=True)
    )

    # Freeze panes (freeze header)
    ws.freeze_panes = "A5"

    auto_fit_columns(ws)
    print("  ✓ Sheet 1: Training Data (with auto-filters & conditional formatting)")
    return last_row


# ═══════════════════════════════════════════════════
# SHEET 2: Employee Lookup (VLOOKUP)
# ═══════════════════════════════════════════════════
def create_lookup_sheet(wb, df, data_last_row):
    """Create a VLOOKUP-based employee lookup sheet."""
    ws = wb.create_sheet("Employee Lookup")
    ws.sheet_properties.tabColor = "2E75B6"

    # Title
    ws.merge_cells("A1:G1")
    ws["A1"] = "EMPLOYEE TRAINING LOOKUP (VLOOKUP)"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:G2")
    ws["A2"] = "Enter an Employee ID below to look up their details using VLOOKUP formulas"
    ws["A2"].font = Font(name="Calibri", italic=True, size=10, color="666666")
    ws["A2"].alignment = Alignment(horizontal="center")

    # ─── Lookup Input Section ───
    ws["A4"] = "🔍 LOOKUP INPUT"
    ws["A4"].font = Font(name="Calibri", bold=True, size=12, color=DARK_BLUE)

    ws["A5"] = "Enter Employee ID:"
    ws["A5"].font = BOLD_FONT
    ws["B5"] = 1303  # Default lookup value
    ws["B5"].font = Font(name="Calibri", bold=True, size=14, color="CC0000")
    ws["B5"].fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
    ws["B5"].border = Border(
        left=Side(style="thick", color="CC0000"),
        right=Side(style="thick", color="CC0000"),
        top=Side(style="thick", color="CC0000"),
        bottom=Side(style="thick", color="CC0000"),
    )

    # ─── VLOOKUP Results ───
    ws["A7"] = "📋 VLOOKUP RESULTS"
    ws["A7"].font = Font(name="Calibri", bold=True, size=12, color=DARK_BLUE)

    lookup_fields = [
        ("Employee Name", 2, ""),
        ("Department", 3, ""),
    ]

    headers_row = 8
    ws.cell(row=headers_row, column=1, value="Field")
    ws.cell(row=headers_row, column=2, value="Formula Used")
    ws.cell(row=headers_row, column=3, value="Result")
    style_header_row(ws, headers_row, 3)

    # VLOOKUP for Employee Name
    r = 9
    ws.cell(row=r, column=1, value="Employee Name")
    ws.cell(row=r, column=1).font = BOLD_FONT
    ws.cell(row=r, column=2, value='=FORMULATEXT(C9)')
    ws.cell(row=r, column=2).font = Font(name="Consolas", size=9, color="666666")
    ws.cell(row=r, column=3).value = f'=VLOOKUP(B5,\'Training Data\'!A5:M{data_last_row},2,FALSE)'
    ws.cell(row=r, column=3).font = Font(name="Calibri", bold=True, size=11)

    # VLOOKUP for Department
    r = 10
    ws.cell(row=r, column=1, value="Department")
    ws.cell(row=r, column=1).font = BOLD_FONT
    ws.cell(row=r, column=2, value='=FORMULATEXT(C10)')
    ws.cell(row=r, column=2).font = Font(name="Consolas", size=9, color="666666")
    ws.cell(row=r, column=3).value = f'=VLOOKUP(B5,\'Training Data\'!A5:M{data_last_row},4,FALSE)'
    ws.cell(row=r, column=3).font = Font(name="Calibri", bold=True, size=11)

    # VLOOKUP for Status
    r = 11
    ws.cell(row=r, column=1, value="Training Status")
    ws.cell(row=r, column=1).font = BOLD_FONT
    ws.cell(row=r, column=2, value='=FORMULATEXT(C11)')
    ws.cell(row=r, column=2).font = Font(name="Consolas", size=9, color="666666")
    ws.cell(row=r, column=3).value = f'=VLOOKUP(B5,\'Training Data\'!A5:M{data_last_row},8,FALSE)'
    ws.cell(row=r, column=3).font = Font(name="Calibri", bold=True, size=11)

    # VLOOKUP for Overdue
    r = 12
    ws.cell(row=r, column=1, value="Is Overdue?")
    ws.cell(row=r, column=1).font = BOLD_FONT
    ws.cell(row=r, column=2, value='=FORMULATEXT(C12)')
    ws.cell(row=r, column=2).font = Font(name="Consolas", size=9, color="666666")
    ws.cell(row=r, column=3).value = f'=VLOOKUP(B5,\'Training Data\'!A5:M{data_last_row},10,FALSE)'
    ws.cell(row=r, column=3).font = Font(name="Calibri", bold=True, size=11)

    style_data_rows(ws, 9, 12, 3)

    # Conditional formatting on results
    ws.conditional_formatting.add(
        "C11",
        CellIsRule(operator="equal", formula=['"Completed"'],
                   fill=PatternFill(bgColor=GREEN), font=Font(color=GREEN_FONT, bold=True))
    )
    ws.conditional_formatting.add(
        "C11",
        CellIsRule(operator="equal", formula=['"Pending"'],
                   fill=PatternFill(bgColor=YELLOW), font=Font(color=YELLOW_FONT, bold=True))
    )
    ws.conditional_formatting.add(
        "C12",
        CellIsRule(operator="equal", formula=['"Yes"'],
                   fill=PatternFill(bgColor=RED), font=Font(color=RED_FONT, bold=True))
    )

    # ─── COUNTIF Summary for the looked-up employee ───
    ws["A14"] = "📊 EMPLOYEE TRAINING SUMMARY (COUNTIF)"
    ws["A14"].font = Font(name="Calibri", bold=True, size=12, color=DARK_BLUE)

    summary_headers = ["Metric", "Formula", "Value"]
    for col, h in enumerate(summary_headers, 1):
        ws.cell(row=15, column=col, value=h)
    style_header_row(ws, 15, 3)

    # Total trainings for this employee
    r = 16
    ws.cell(row=r, column=1, value="Total Trainings Assigned")
    ws.cell(row=r, column=1).font = BOLD_FONT
    ws.cell(row=r, column=2, value='=FORMULATEXT(C16)')
    ws.cell(row=r, column=2).font = Font(name="Consolas", size=9, color="666666")
    ws.cell(row=r, column=3).value = f"=COUNTIF('Training Data'!A5:A{data_last_row},B5)"

    r = 17
    ws.cell(row=r, column=1, value="Completed Trainings")
    ws.cell(row=r, column=1).font = BOLD_FONT
    ws.cell(row=r, column=2, value='=FORMULATEXT(C17)')
    ws.cell(row=r, column=2).font = Font(name="Consolas", size=9, color="666666")
    ws.cell(row=r, column=3).value = f"=COUNTIFS('Training Data'!A5:A{data_last_row},B5,'Training Data'!H5:H{data_last_row},\"Completed\")"

    r = 18
    ws.cell(row=r, column=1, value="Pending Trainings")
    ws.cell(row=r, column=1).font = BOLD_FONT
    ws.cell(row=r, column=2, value='=FORMULATEXT(C18)')
    ws.cell(row=r, column=2).font = Font(name="Consolas", size=9, color="666666")
    ws.cell(row=r, column=3).value = f"=COUNTIFS('Training Data'!A5:A{data_last_row},B5,'Training Data'!H5:H{data_last_row},\"Pending\")"

    r = 19
    ws.cell(row=r, column=1, value="Overdue Trainings")
    ws.cell(row=r, column=1).font = BOLD_FONT
    ws.cell(row=r, column=2, value='=FORMULATEXT(C19)')
    ws.cell(row=r, column=2).font = Font(name="Consolas", size=9, color="666666")
    ws.cell(row=r, column=3).value = f"=COUNTIFS('Training Data'!A5:A{data_last_row},B5,'Training Data'!J5:J{data_last_row},\"Yes\")"

    r = 20
    ws.cell(row=r, column=1, value="Completion Rate (%)")
    ws.cell(row=r, column=1).font = BOLD_FONT
    ws.cell(row=r, column=2, value='=FORMULATEXT(C20)')
    ws.cell(row=r, column=2).font = Font(name="Consolas", size=9, color="666666")
    ws.cell(row=r, column=3).value = "=ROUND(C17/C16*100,1)"
    ws.cell(row=r, column=3).font = Font(name="Calibri", bold=True, size=12, color=DARK_BLUE)

    style_data_rows(ws, 16, 20, 3)

    # ─── Unique Employees Reference Table (for VLOOKUP source) ───
    ws["E4"] = "📋 EMPLOYEE DIRECTORY"
    ws["E4"].font = Font(name="Calibri", bold=True, size=12, color=DARK_BLUE)

    unique_emps = df.drop_duplicates(subset='employee_id')[['employee_id', 'employee_name', 'department']].sort_values('employee_id')

    dir_headers = ["Emp ID", "Name", "Department"]
    for col, h in enumerate(dir_headers, 5):
        ws.cell(row=5, column=col, value=h)
    style_header_row(ws, 5, 7)

    for i, (_, emp) in enumerate(unique_emps.head(30).iterrows()):
        r = i + 6
        ws.cell(row=r, column=5, value=emp['employee_id'])
        ws.cell(row=r, column=6, value=emp['employee_name'])
        ws.cell(row=r, column=7, value=emp['department'])
    style_data_rows(ws, 6, 6 + min(30, len(unique_emps)) - 1, 7)

    # Column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 3
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 22
    ws.column_dimensions['G'].width = 15

    ws.freeze_panes = "A4"
    print("  ✓ Sheet 2: Employee Lookup (VLOOKUP & COUNTIFS formulas)")


# ═══════════════════════════════════════════════════
# SHEET 3: Pivot Table — Department Summary
# ═══════════════════════════════════════════════════
def create_dept_pivot_sheet(wb, df, data_last_row):
    """Create department pivot table with formulas."""
    ws = wb.create_sheet("Pivot - Departments")
    ws.sheet_properties.tabColor = "548235"

    ws.merge_cells("A1:H1")
    ws["A1"] = "PIVOT TABLE — DEPARTMENT PERFORMANCE SUMMARY"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:H2")
    ws["A2"] = "Built using COUNTIFS, AVERAGEIFS, and calculated columns"
    ws["A2"].font = Font(name="Calibri", italic=True, size=10, color="666666")
    ws["A2"].alignment = Alignment(horizontal="center")

    # Headers
    headers = [
        "Department", "Total Trainings", "Completed", "Pending",
        "Overdue", "Completion Rate (%)", "Overdue Rate (%)", "Avg Completion Days"
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=4, column=col, value=h)
    style_header_row(ws, 4, len(headers))

    # Department list
    departments = sorted(df['department'].unique())
    data_ref = "'Training Data'"

    for i, dept in enumerate(departments):
        r = i + 5

        # Department name
        ws.cell(row=r, column=1, value=dept)
        ws.cell(row=r, column=1).font = BOLD_FONT

        # Total: =COUNTIF(dept_range, dept)
        ws.cell(row=r, column=2).value = f"=COUNTIF({data_ref}!C5:C{data_last_row},A{r})"

        # Completed: =COUNTIFS(dept_range, dept, status_range, "Completed")
        ws.cell(row=r, column=3).value = f'=COUNTIFS({data_ref}!C5:C{data_last_row},A{r},{data_ref}!H5:H{data_last_row},"Completed")'

        # Pending
        ws.cell(row=r, column=4).value = f'=COUNTIFS({data_ref}!C5:C{data_last_row},A{r},{data_ref}!H5:H{data_last_row},"Pending")'

        # Overdue
        ws.cell(row=r, column=5).value = f'=COUNTIFS({data_ref}!C5:C{data_last_row},A{r},{data_ref}!J5:J{data_last_row},"Yes")'

        # Completion Rate = Completed / Total * 100
        ws.cell(row=r, column=6).value = f"=ROUND(C{r}/B{r}*100,1)"
        ws.cell(row=r, column=6).number_format = '0.0'

        # Overdue Rate = Overdue / Pending * 100
        ws.cell(row=r, column=7).value = f'=IF(D{r}=0,0,ROUND(E{r}/D{r}*100,1))'
        ws.cell(row=r, column=7).number_format = '0.0'

        # Avg Completion Days
        ws.cell(row=r, column=8).value = f'=ROUND(AVERAGEIFS({data_ref}!I5:I{data_last_row},{data_ref}!C5:C{data_last_row},A{r},{data_ref}!H5:H{data_last_row},"Completed"),1)'
        ws.cell(row=r, column=8).number_format = '0.0'

    last_dept_row = 4 + len(departments)
    style_data_rows(ws, 5, last_dept_row, len(headers))

    # Grand Total row
    total_row = last_dept_row + 1
    ws.cell(row=total_row, column=1, value="GRAND TOTAL")
    ws.cell(row=total_row, column=1).font = Font(name="Calibri", bold=True, size=11, color=WHITE)
    for col in range(1, len(headers) + 1):
        ws.cell(row=total_row, column=col).fill = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid")
        ws.cell(row=total_row, column=col).font = Font(name="Calibri", bold=True, color=WHITE, size=10)
        ws.cell(row=total_row, column=col).border = THIN_BORDER

    ws.cell(row=total_row, column=2).value = f"=SUM(B5:B{last_dept_row})"
    ws.cell(row=total_row, column=3).value = f"=SUM(C5:C{last_dept_row})"
    ws.cell(row=total_row, column=4).value = f"=SUM(D5:D{last_dept_row})"
    ws.cell(row=total_row, column=5).value = f"=SUM(E5:E{last_dept_row})"
    ws.cell(row=total_row, column=6).value = f"=ROUND(C{total_row}/B{total_row}*100,1)"
    ws.cell(row=total_row, column=7).value = f"=IF(D{total_row}=0,0,ROUND(E{total_row}/D{total_row}*100,1))"
    ws.cell(row=total_row, column=8).value = f"=ROUND(AVERAGE(H5:H{last_dept_row}),1)"

    # Conditional formatting: Completion Rate color scale
    ws.conditional_formatting.add(
        f"F5:F{last_dept_row}",
        ColorScaleRule(
            start_type="min", start_color="F8696B",
            mid_type="percentile", mid_value=50, mid_color="FFEB84",
            end_type="max", end_color="63BE7B"
        )
    )

    # Data bars on Total column
    ws.conditional_formatting.add(
        f"B5:B{last_dept_row}",
        DataBarRule(start_type="min", end_type="max", color="2E75B6", showValue=True)
    )

    # ─── Bar Chart: Department Completion ───
    chart = BarChart()
    chart.type = "col"
    chart.title = "Department Completion: Completed vs Pending"
    chart.y_axis.title = "Count"
    chart.x_axis.title = "Department"
    chart.style = 10

    cats = Reference(ws, min_col=1, min_row=5, max_row=last_dept_row)
    completed_data = Reference(ws, min_col=3, min_row=4, max_row=last_dept_row)
    pending_data = Reference(ws, min_col=4, min_row=4, max_row=last_dept_row)

    chart.add_data(completed_data, titles_from_data=True)
    chart.add_data(pending_data, titles_from_data=True)
    chart.set_categories(cats)
    chart.series[0].graphicalProperties.solidFill = "548235"
    chart.series[1].graphicalProperties.solidFill = "FFC000"
    chart.width = 22
    chart.height = 14

    ws.add_chart(chart, f"A{total_row + 3}")

    auto_fit_columns(ws, min_width=15)
    ws.freeze_panes = "A5"
    print("  ✓ Sheet 3: Pivot Table — Department Summary (COUNTIFS formulas + chart)")


# ═══════════════════════════════════════════════════
# SHEET 4: Pivot Table — Course Summary
# ═══════════════════════════════════════════════════
def create_course_pivot_sheet(wb, df, data_last_row):
    """Create course pivot table with formulas."""
    ws = wb.create_sheet("Pivot - Courses")
    ws.sheet_properties.tabColor = "BF8F00"

    ws.merge_cells("A1:G1")
    ws["A1"] = "PIVOT TABLE — COURSE PERFORMANCE SUMMARY"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="center")

    headers = [
        "Course Name", "Total Assigned", "Completed", "Pending",
        "Completion Rate (%)", "Avg Days to Complete",
        "Status Flag"
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=h)
    style_header_row(ws, 3, len(headers))

    courses = sorted(df['course_name'].unique())
    data_ref = "'Training Data'"

    for i, course in enumerate(courses):
        r = i + 4
        ws.cell(row=r, column=1, value=course)
        ws.cell(row=r, column=1).font = BOLD_FONT

        ws.cell(row=r, column=2).value = f'=COUNTIF({data_ref}!D5:D{data_last_row},A{r})'
        ws.cell(row=r, column=3).value = f'=COUNTIFS({data_ref}!D5:D{data_last_row},A{r},{data_ref}!H5:H{data_last_row},"Completed")'
        ws.cell(row=r, column=4).value = f'=COUNTIFS({data_ref}!D5:D{data_last_row},A{r},{data_ref}!H5:H{data_last_row},"Pending")'
        ws.cell(row=r, column=5).value = f"=ROUND(C{r}/B{r}*100,1)"
        ws.cell(row=r, column=5).number_format = '0.0'
        ws.cell(row=r, column=6).value = f'=ROUND(AVERAGEIFS({data_ref}!I5:I{data_last_row},{data_ref}!D5:D{data_last_row},A{r},{data_ref}!H5:H{data_last_row},"Completed"),1)'
        ws.cell(row=r, column=6).number_format = '0.0'

        # IF formula: Flag courses below target
        ws.cell(row=r, column=7).value = f'=IF(E{r}>=75,"✅ On Track",IF(E{r}>=60,"⚠️ Needs Attention","🔴 Critical"))'

    last_row = 3 + len(courses)
    style_data_rows(ws, 4, last_row, len(headers))

    # Conditional formatting on completion rate
    ws.conditional_formatting.add(
        f"E4:E{last_row}",
        ColorScaleRule(
            start_type="min", start_color="F8696B",
            mid_type="percentile", mid_value=50, mid_color="FFEB84",
            end_type="max", end_color="63BE7B"
        )
    )

    # Chart
    chart = BarChart()
    chart.type = "bar"
    chart.title = "Course Completion Rates (%)"
    chart.x_axis.title = "Completion Rate (%)"
    chart.style = 10

    cats = Reference(ws, min_col=1, min_row=4, max_row=last_row)
    rates = Reference(ws, min_col=5, min_row=3, max_row=last_row)
    chart.add_data(rates, titles_from_data=True)
    chart.set_categories(cats)
    chart.series[0].graphicalProperties.solidFill = "2E75B6"
    chart.width = 22
    chart.height = 14

    ws.add_chart(chart, f"A{last_row + 3}")

    auto_fit_columns(ws, min_width=15)
    ws.freeze_panes = "A4"
    print("  ✓ Sheet 4: Pivot Table — Course Summary (with IF status flags + chart)")


# ═══════════════════════════════════════════════════
# SHEET 5: Overdue Tracker (IF formulas)
# ═══════════════════════════════════════════════════
def create_overdue_sheet(wb, df, data_last_row):
    """Create overdue tracking sheet with IF formulas and conditional formatting."""
    ws = wb.create_sheet("Overdue Tracker")
    ws.sheet_properties.tabColor = "C00000"

    ws.merge_cells("A1:H1")
    ws["A1"] = "⚠️ OVERDUE TRAINING TRACKER"
    ws["A1"].font = Font(name="Calibri", bold=True, size=14, color="C00000")
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:H2")
    ws["A2"] = "Employees with pending trainings past their due date — Requires immediate action"
    ws["A2"].font = Font(name="Calibri", italic=True, size=10, color="666666")
    ws["A2"].alignment = Alignment(horizontal="center")

    # Get overdue records
    reference_date = pd.Timestamp('2026-03-31')
    overdue_df = df[df['is_overdue'] == True].copy()
    overdue_df['days_overdue'] = (reference_date - overdue_df['due_date']).dt.days

    # Summary by employee
    emp_overdue = overdue_df.groupby(['employee_id', 'employee_name', 'department']).agg(
        overdue_courses=('course_name', 'count'),
        courses_list=('course_name', lambda x: ', '.join(sorted(x))),
        max_days_overdue=('days_overdue', 'max'),
        avg_days_overdue=('days_overdue', 'mean'),
    ).reset_index()
    emp_overdue['avg_days_overdue'] = emp_overdue['avg_days_overdue'].round(0).astype(int)
    emp_overdue = emp_overdue.sort_values('overdue_courses', ascending=False)

    # Headers
    headers = [
        "Employee ID", "Employee Name", "Department", "Overdue Courses",
        "Course Names", "Max Days Overdue", "Avg Days Overdue",
        "Priority Level"
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=4, column=col, value=h)
    style_header_row(ws, 4, len(headers))

    # Data + IF formula for priority
    for i, (_, row) in enumerate(emp_overdue.iterrows()):
        r = i + 5
        ws.cell(row=r, column=1, value=row['employee_id'])
        ws.cell(row=r, column=2, value=row['employee_name'])
        ws.cell(row=r, column=3, value=row['department'])
        ws.cell(row=r, column=4, value=row['overdue_courses'])
        ws.cell(row=r, column=5, value=row['courses_list'])
        ws.cell(row=r, column=6, value=row['max_days_overdue'])
        ws.cell(row=r, column=7, value=row['avg_days_overdue'])

        # IF formula for priority level
        ws.cell(row=r, column=8).value = f'=IF(D{r}>=8,"🔴 CRITICAL",IF(D{r}>=5,"🟡 HIGH",IF(D{r}>=3,"🟠 MEDIUM","🟢 LOW")))'

    last_row = 4 + len(emp_overdue)
    style_data_rows(ws, 5, last_row, len(headers))

    # Auto-filter
    ws.auto_filter.ref = f"A4:H{last_row}"

    # Conditional formatting on overdue courses count
    ws.conditional_formatting.add(
        f"D5:D{last_row}",
        ColorScaleRule(
            start_type="min", start_color="63BE7B",
            mid_type="percentile", mid_value=50, mid_color="FFEB84",
            end_type="max", end_color="F8696B"
        )
    )

    # Conditional formatting on days overdue
    ws.conditional_formatting.add(
        f"F5:F{last_row}",
        DataBarRule(start_type="min", end_type="max", color="C00000", showValue=True)
    )

    # KPI summary at top-right
    ws["J4"] = "OVERDUE SUMMARY"
    ws["J4"].font = Font(name="Calibri", bold=True, size=12, color="C00000")

    summary_items = [
        ("Total Overdue Trainings", len(overdue_df)),
        ("Employees Affected", len(emp_overdue)),
        ("Avg Days Overdue", int(overdue_df['days_overdue'].mean())),
        ("Max Days Overdue", int(overdue_df['days_overdue'].max())),
        ("Critical (8+ courses)", len(emp_overdue[emp_overdue['overdue_courses'] >= 8])),
    ]
    for i, (label, value) in enumerate(summary_items):
        r = 5 + i
        ws.cell(row=r, column=10, value=label)
        ws.cell(row=r, column=10).font = BOLD_FONT
        ws.cell(row=r, column=11, value=value)
        ws.cell(row=r, column=11).font = Font(name="Calibri", bold=True, size=12, color="C00000")
        ws.cell(row=r, column=11).alignment = Alignment(horizontal="center")

    auto_fit_columns(ws, min_width=12, max_width=50)
    ws.column_dimensions['E'].width = 45
    ws.freeze_panes = "A5"
    print("  ✓ Sheet 5: Overdue Tracker (IF formulas + auto-filters + conditional formatting)")


# ═══════════════════════════════════════════════════
# SHEET 6: KPI Dashboard
# ═══════════════════════════════════════════════════
def create_kpi_dashboard_sheet(wb, df, data_last_row):
    """Create a KPI dashboard sheet with formula-driven cards."""
    ws = wb.create_sheet("KPI Dashboard")
    ws.sheet_properties.tabColor = "7030A0"

    # Move this sheet to be first (after Training Data)
    wb.move_sheet(ws, offset=-4)

    ws.merge_cells("A1:L1")
    ws["A1"] = "📊 LMS TRAINING DASHBOARD — KEY PERFORMANCE INDICATORS"
    ws["A1"].font = Font(name="Calibri", bold=True, size=16, color=DARK_BLUE)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:L2")
    ws["A2"] = "All values are formula-driven — connected to the Training Data sheet"
    ws["A2"].font = Font(name="Calibri", italic=True, size=10, color="666666")
    ws["A2"].alignment = Alignment(horizontal="center")

    data_ref = "'Training Data'"

    # ─── KPI Cards (Row 4-7) ───
    kpis = [
        ("Total Records", f"=COUNTA({data_ref}!A5:A{data_last_row})", "A"),
        ("Unique Employees", f"=SUMPRODUCT(1/COUNTIF({data_ref}!A5:A{data_last_row},{data_ref}!A5:A{data_last_row}))", "C"),
        ("Completed", f'=COUNTIF({data_ref}!H5:H{data_last_row},"Completed")', "E"),
        ("Pending", f'=COUNTIF({data_ref}!H5:H{data_last_row},"Pending")', "G"),
        ("Overdue", f'=COUNTIF({data_ref}!J5:J{data_last_row},"Yes")', "I"),
        ("Completion Rate", f'=ROUND(COUNTIF({data_ref}!H5:H{data_last_row},"Completed")/COUNTA({data_ref}!H5:H{data_last_row})*100,1)&"%"', "K"),
    ]

    kpi_colors = ["1F4E79", "2E75B6", "548235", "BF8F00", "C00000", "7030A0"]

    for i, (label, formula, col_letter) in enumerate(kpis):
        col_num = ord(col_letter) - ord('A') + 1

        # Merge 2 columns for each KPI card
        ws.merge_cells(f"{col_letter}4:{get_column_letter(col_num+1)}4")
        ws.merge_cells(f"{col_letter}5:{get_column_letter(col_num+1)}5")
        ws.merge_cells(f"{col_letter}6:{get_column_letter(col_num+1)}6")

        # Label
        ws[f"{col_letter}4"] = label.upper()
        ws[f"{col_letter}4"].font = Font(name="Calibri", bold=True, size=9, color=WHITE)
        ws[f"{col_letter}4"].fill = PatternFill(start_color=kpi_colors[i], end_color=kpi_colors[i], fill_type="solid")
        ws[f"{col_letter}4"].alignment = Alignment(horizontal="center")
        ws[f"{get_column_letter(col_num+1)}4"].fill = PatternFill(start_color=kpi_colors[i], end_color=kpi_colors[i], fill_type="solid")

        # Value (formula)
        ws[f"{col_letter}5"] = formula
        ws[f"{col_letter}5"].font = Font(name="Calibri", bold=True, size=22, color=kpi_colors[i])
        ws[f"{col_letter}5"].alignment = Alignment(horizontal="center", vertical="center")

        # Bottom border
        ws[f"{col_letter}6"] = ""
        for c in range(col_num, col_num + 2):
            ws.cell(row=6, column=c).border = Border(bottom=Side(style="thick", color=kpi_colors[i]))

    # ─── Formula Reference Table ───
    ws["A8"] = "📋 FORMULAS USED IN THIS DASHBOARD"
    ws["A8"].font = Font(name="Calibri", bold=True, size=12, color=DARK_BLUE)

    formula_headers = ["Formula Type", "Purpose", "Example"]
    for col, h in enumerate(formula_headers, 1):
        ws.cell(row=9, column=col, value=h)
    style_header_row(ws, 9, 3)

    formulas_used = [
        ("VLOOKUP", "Look up employee details by ID", '=VLOOKUP(ID, range, col, FALSE)'),
        ("COUNTIF", "Count records matching criteria", '=COUNTIF(range, "Completed")'),
        ("COUNTIFS", "Count with multiple conditions", '=COUNTIFS(dept,"HR", status,"Completed")'),
        ("AVERAGEIFS", "Average with conditions", '=AVERAGEIFS(days, status, "Completed")'),
        ("IF / Nested IF", "Classify priority levels", '=IF(count>=8,"Critical",IF(...))'),
        ("ROUND", "Round calculated percentages", '=ROUND(value, 1)'),
        ("SUMPRODUCT", "Count unique values", '=SUMPRODUCT(1/COUNTIF(...))'),
        ("FORMULATEXT", "Display formula as text", '=FORMULATEXT(C9)'),
        ("Auto-Filters", "Filter data by any column", "Applied on Training Data sheet"),
        ("Conditional Formatting", "Visual alerts for status", "Red/Yellow/Green on status columns"),
        ("Data Bars", "In-cell bar charts", "Applied on numeric columns"),
        ("Color Scales", "Gradient highlighting", "Applied on completion rate columns"),
    ]

    for i, (ftype, purpose, example) in enumerate(formulas_used):
        r = 10 + i
        ws.cell(row=r, column=1, value=ftype)
        ws.cell(row=r, column=1).font = BOLD_FONT
        ws.cell(row=r, column=2, value=purpose)
        ws.cell(row=r, column=3, value=example)
        ws.cell(row=r, column=3).font = Font(name="Consolas", size=9, color="666666")

    style_data_rows(ws, 10, 10 + len(formulas_used) - 1, 3)

    # Column widths
    for col_letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']:
        ws.column_dimensions[col_letter].width = 15
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 40

    # Row heights for KPI cards
    ws.row_dimensions[4].height = 25
    ws.row_dimensions[5].height = 45
    ws.row_dimensions[6].height = 8

    print("  ✓ Sheet 6: KPI Dashboard (formula-driven cards + formula reference)")


# ═══════════════════════════════════════════════════
# SHEET 7: Monthly Trends
# ═══════════════════════════════════════════════════
def create_monthly_sheet(wb, df, data_last_row):
    """Create monthly trends sheet with chart."""
    ws = wb.create_sheet("Monthly Trends")
    ws.sheet_properties.tabColor = "00B0F0"

    ws.merge_cells("A1:F1")
    ws["A1"] = "MONTHLY TRAINING TRENDS"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="center")

    headers = ["Month", "Total Assigned", "Completed", "Pending", "Completion Rate (%)", "Status"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=3, column=col, value=h)
    style_header_row(ws, 3, len(headers))

    monthly = df.groupby('assigned_month').agg(
        total=('status', 'count'),
        completed=('status', lambda x: (x == 'Completed').sum()),
        pending=('status', lambda x: (x == 'Pending').sum()),
    ).reset_index().sort_values('assigned_month')
    monthly['rate'] = round(monthly['completed'] / monthly['total'] * 100, 1)

    for i, (_, row) in enumerate(monthly.iterrows()):
        r = i + 4
        ws.cell(row=r, column=1, value=row['assigned_month'])
        ws.cell(row=r, column=2, value=row['total'])
        ws.cell(row=r, column=3, value=row['completed'])
        ws.cell(row=r, column=4, value=row['pending'])
        ws.cell(row=r, column=5, value=row['rate'])
        ws.cell(row=r, column=5).number_format = '0.0'
        # IF formula for trend status
        ws.cell(row=r, column=6).value = f'=IF(E{r}>=70,"✅ Good",IF(E{r}>=65,"⚠️ Average","🔴 Below Target"))'

    last_row = 3 + len(monthly)
    style_data_rows(ws, 4, last_row, len(headers))

    # Color scale on completion rate
    ws.conditional_formatting.add(
        f"E4:E{last_row}",
        ColorScaleRule(
            start_type="min", start_color="F8696B",
            mid_type="percentile", mid_value=50, mid_color="FFEB84",
            end_type="max", end_color="63BE7B"
        )
    )

    # Line Chart
    chart = LineChart()
    chart.title = "Monthly Training Trends"
    chart.y_axis.title = "Count"
    chart.x_axis.title = "Month"
    chart.style = 10
    chart.width = 22
    chart.height = 14

    cats = Reference(ws, min_col=1, min_row=4, max_row=last_row)
    total_data = Reference(ws, min_col=2, min_row=3, max_row=last_row)
    comp_data = Reference(ws, min_col=3, min_row=3, max_row=last_row)
    pend_data = Reference(ws, min_col=4, min_row=3, max_row=last_row)

    chart.add_data(total_data, titles_from_data=True)
    chart.add_data(comp_data, titles_from_data=True)
    chart.add_data(pend_data, titles_from_data=True)
    chart.set_categories(cats)

    chart.series[0].graphicalProperties.line.solidFill = "2E75B6"
    chart.series[1].graphicalProperties.line.solidFill = "548235"
    chart.series[2].graphicalProperties.line.solidFill = "FFC000"

    ws.add_chart(chart, f"A{last_row + 3}")

    auto_fit_columns(ws, min_width=15)
    print("  ✓ Sheet 7: Monthly Trends (with line chart + IF status flags)")


# ═══════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════
def main():
    print("=" * 60)
    print("  LMS Excel Dashboard Builder")
    print("=" * 60)

    base_dir = os.path.dirname(os.path.dirname(__file__))
    clean_path = os.path.join(base_dir, "data", "lms_cleaned.csv")
    output_path = os.path.join(base_dir, "reports", "dashboard.xlsx")
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    # Load data
    df = pd.read_csv(clean_path, parse_dates=['assigned_date', 'due_date', 'completion_date'])
    print(f"\n✓ Loaded {len(df)} records")

    # Create workbook
    wb = Workbook()
    print(f"\nBuilding Excel sheets...\n")

    # Build sheets
    data_last_row = create_data_sheet(wb, df)
    create_kpi_dashboard_sheet(wb, df, data_last_row)
    create_lookup_sheet(wb, df, data_last_row)
    create_dept_pivot_sheet(wb, df, data_last_row)
    create_course_pivot_sheet(wb, df, data_last_row)
    create_overdue_sheet(wb, df, data_last_row)
    create_monthly_sheet(wb, df, data_last_row)

    # Save
    wb.save(output_path)
    print(f"\n✅ Excel dashboard saved to: {output_path}")
    print(f"\n📋 Sheets created:")
    print(f"   1. KPI Dashboard      — Formula-driven KPI cards")
    print(f"   2. Training Data      — Full dataset with auto-filters & conditional formatting")
    print(f"   3. Employee Lookup    — VLOOKUP & COUNTIFS formulas")
    print(f"   4. Pivot - Departments — COUNTIFS-based pivot table + bar chart")
    print(f"   5. Pivot - Courses    — Course pivot table + IF status flags + chart")
    print(f"   6. Overdue Tracker    — Nested IF priority levels + auto-filters")
    print(f"   7. Monthly Trends     — Trend data + line chart + IF formulas")
    print(f"\n💡 Open in Excel to see all formulas, filters, and conditional formatting!")


if __name__ == "__main__":
    main()
